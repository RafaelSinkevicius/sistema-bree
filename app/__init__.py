from flask import Flask, render_template, request, redirect, url_for, flash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, timezone, timedelta
from flask import request
import pandas as pd
import io
from flask import send_file
from sqlalchemy import extract
from sqlalchemy.orm import joinedload, load_only
from sqlalchemy.exc import IntegrityError
from flask import session
from werkzeug.security import check_password_hash, generate_password_hash
from functools import wraps
from decimal import Decimal
from statistics import median
import logging

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "usuario_id" not in session:
            flash("Você precisa estar logado para acessar esta página.")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("usuario_tipo") != "admin":
            flash("Apenas administradores podem acessar esta página.")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated_function

def agora_brasil():
    return datetime.now(tz=timezone(timedelta(hours=-3)))


def importar_contratos_de_planilha(caminho_ou_buffer):
    """
    Lê uma planilha Excel e importa contratos e vendedores.
    `caminho_ou_buffer` pode ser um caminho de arquivo ou um objeto file-like (upload).
    """
    df = pd.read_excel(caminho_ou_buffer)

    total_ok = 0
    total_duplicados = 0
    total_erros = 0
    mensagens = []

    for index, row in df.iterrows():
        try:
            # Processa o vendedor
            nome_vendedor = str(row['VENDEDOR']).strip()
            vendedor = Vendedor.query.filter_by(nome=nome_vendedor).first()
            if not vendedor:
                vendedor = Vendedor(
                    nome=nome_vendedor,
                    cpf_cnpj='00000000000',
                    celular='',
                    email=''
                )
                db.session.add(vendedor)
                db.session.commit()

            # Trata o campo de Mês de Cancelamento
            mes_cancelamento = pd.to_datetime(
                row['Mês de Cancelamento (se aplicável)'],
                errors='coerce'
            )
            mes_cancelamento = None if pd.isna(mes_cancelamento) else mes_cancelamento

            contrato = Contrato(
                proposta=row['PROPOSTA'],
                data_checagem=pd.to_datetime(row['DATA DE CHECAGEM'], errors='coerce'),
                razao_social=row['RAZÃO SOCIAL/NOME'],
                cnpj_cpf=row['CNPJ/CPF'],
                celular=row['CELULAR'],
                email=row['E-MAIL'],
                atividade_economica=row['ATIVIDADE ECONÔMICA'],
                cidade=row['CIDADE'],
                nome_plano=row['NOME DO PLANO'],
                data_vigencia=pd.to_datetime(row['DATA DE VIGÊNCIA'], errors='coerce'),
                vidas=int(row['VIDAS']) if pd.notna(row['VIDAS']) else None,
                valor_parcela=float(row['VALOR DA PARCELA']) if pd.notna(row['VALOR DA PARCELA']) else None,
                parcela_atual=int(row['PARCELA ATUAL']) if pd.notna(row['PARCELA ATUAL']) else None,
                status=row['STATUS'],
                mes_cancelamento=mes_cancelamento,
                verificado=(
                    str(row['VERIFICADO?']).strip().lower() == 'sim'
                ) if pd.notna(row['VERIFICADO?']) else False,
                contrato=row['CONTRATO'],
                vendedor_id=vendedor.id
            )

            db.session.add(contrato)
            db.session.commit()
            total_ok += 1

        except IntegrityError:
            db.session.rollback()
            total_duplicados += 1
            mensagens.append(f"Contrato {row['CONTRATO']} já existe. Pulado.")
        except Exception as e:
            db.session.rollback()
            total_erros += 1
            mensagens.append(f"Erro ao importar contrato {row.get('CONTRATO', index)}: {e}")

    resumo = (
        f"Importação concluída. "
        f"{total_ok} contratos importados, "
        f"{total_duplicados} duplicados, "
        f"{total_erros} com erro."
    )
    logging.info(resumo)
    return resumo, mensagens


def sobrepor_status_de_planilha(caminho_ou_buffer):
    """
    Lê uma planilha Excel e SOBREPÕE apenas o STATUS dos contratos existentes.
    A planilha precisa ter pelo menos as colunas: CONTRATO e STATUS
    
    Uso: Permite que o admin force a mudança de status de contratos em massa,
    por exemplo, para "ressuscitar" contratos cancelados ou corrigir status incorretos.
    """
    df = pd.read_excel(caminho_ou_buffer)

    total_atualizados = 0
    total_nao_encontrados = 0
    total_erros = 0
    mensagens = []

    # Verifica se as colunas obrigatórias existem
    if 'CONTRATO' not in df.columns or 'STATUS' not in df.columns:
        return "❌ Erro: A planilha precisa ter as colunas 'CONTRATO' e 'STATUS'.", []

    for index, row in df.iterrows():
        try:
            numero_contrato = str(row['CONTRATO']).strip()
            novo_status = str(row['STATUS']).strip()

            # Busca o contrato existente
            contrato = Contrato.query.filter_by(contrato=numero_contrato).first()

            if not contrato:
                total_nao_encontrados += 1
                mensagens.append(f"⚠️ Contrato {numero_contrato} não encontrado. Pulado.")
                continue

            # Atualiza apenas o status
            status_antigo = contrato.status
            contrato.status = novo_status
            db.session.commit()
            
            total_atualizados += 1
            mensagens.append(f"✅ Contrato {numero_contrato}: {status_antigo} → {novo_status}")

        except Exception as e:
            db.session.rollback()
            total_erros += 1
            mensagens.append(f"❌ Erro no contrato {row.get('CONTRATO', index)}: {e}")

    resumo = (
        f"Sobreposição concluída. "
        f"{total_atualizados} atualizados, "
        f"{total_nao_encontrados} não encontrados, "
        f"{total_erros} com erro."
    )
    logging.info(f"[SOBREPOR STATUS] {resumo}")
    return resumo, mensagens


app = Flask(__name__)

# AJUSTE PARA PYINSTALLER (Frozen Path)
import sys
import os

import os
import sys
from dotenv import load_dotenv

if getattr(sys, 'frozen', False):
    template_folder = os.path.join(sys._MEIPASS, 'app', 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'app', 'static')
    
    # LOAD ENV ROBUSTO (Correção Definitiva)
    basedir = os.path.dirname(sys.executable)
    load_dotenv(os.path.join(basedir, ".env"))
    
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
else:
    app = Flask(__name__)

# Configuração da Chave Secreta (Cookies/Sessão)
app.secret_key = os.getenv("SECRET_KEY")
if not app.secret_key:
    raise RuntimeError("A variável de ambiente 'SECRET_KEY' não está configurada. Adicione ao .env.")

# Configuração básica de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)

@app.before_request
def checar_login():
    if request.endpoint not in ["login", "static"] and "usuario_id" not in session:
        return redirect("/login")


# Configuração do Banco de Dados (Via Variável de Ambiente)
# A senha fica no arquivo .env (que não vai pro Git)
uri = os.getenv("DATABASE_URL")

if not uri:
    raise RuntimeError("A variável de ambiente 'DATABASE_URL' não está configurada. Crie um arquivo .env com a string de conexão.")

if uri and uri.startswith("postgres://"):
    uri = uri.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = uri
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['JSON_AS_ASCII'] = False

db = SQLAlchemy(app)

# Modelo Contrato
    # Modelo Contrato
class Contrato(db.Model):
    __tablename__ = 'contratos'

    id = db.Column(db.Integer, primary_key=True)
    proposta = db.Column(db.String(50), unique=True, nullable=False)
    data_checagem = db.Column(db.Date)
    razao_social = db.Column(db.String(255))
    cnpj_cpf = db.Column(db.String(20))
    celular = db.Column(db.String(20))
    email = db.Column(db.String(255))
    atividade_economica = db.Column(db.String(255))
    cidade = db.Column(db.String(100))
    nome_plano = db.Column(db.String(255))
    data_vigencia = db.Column(db.Date)
    vidas = db.Column(db.Integer)
    valor_parcela = db.Column(db.Numeric(10,2))
    parcela_atual = db.Column(db.Integer)
    status = db.Column(db.String(50))
    mes_cancelamento = db.Column(db.Date, nullable=True)
    verificado = db.Column(db.Boolean, default=False)
    contrato = db.Column(db.String(50), unique=True)
    vendedor_id = db.Column(db.Integer, db.ForeignKey('vendedores.id'))
    dias_atraso = db.Column(db.Integer, nullable=True)
    
    # NOTA: Cliente Morto NÃO EXISTE MAIS.
    # Contratos com 61+ dias de atraso são Cancelados por Inadimplência automaticamente.
    
    acoes = db.relationship('AcaoCobranca', backref='contrato', lazy=True)
    envio_sms = db.Column(db.Boolean, default=True)
    cliente_critico = db.Column(db.Boolean, default=False)

    vendedor = db.relationship('Vendedor', backref=db.backref('contratos', lazy=True))

# Modelo Vendedor
class Vendedor(db.Model):
    __tablename__ = 'vendedores'
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(255), unique=True)
    cpf_cnpj = db.Column(db.String(20), nullable=True)
    celular = db.Column(db.String(20), nullable=True)
    email = db.Column(db.String(255), nullable=True)

class AcaoCobranca(db.Model):
    __tablename__ = "acoes_cobranca"
    id = db.Column(db.Integer, primary_key=True)
    contrato_id = db.Column(db.Integer, db.ForeignKey("contratos.id"), nullable=False)
    tipo = db.Column(db.String(50))  # Ex: SMS, WhatsApp
    mensagem = db.Column(db.Text)
    dia_atraso = db.Column(db.Integer)
    parcela = db.Column(db.Integer)
    enviada_em = db.Column(db.DateTime)
    status_envio = db.Column(db.String(50))
    usuario = db.Column(db.String(255), nullable=True)



class ResponsavelCobranca(db.Model):
    __tablename__ = "responsaveis_cobranca"
    id = db.Column(db.Integer, primary_key=True)
    contrato_id = db.Column(db.Integer, db.ForeignKey("contratos.id"), nullable=False)
    usuario = db.Column(db.String(255), nullable=False)

class Usuario(db.Model):
    __tablename__ = "usuarios"
    id = db.Column(db.Integer, primary_key=True)
    nome = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    senha = db.Column(db.String(255), nullable=False)
    tipo = db.Column(db.String(20), nullable=False)  # admin ou usuario



# Rota do dashboard inicial (página principal do sistema)
@app.route('/')
def dashboard():
    # Calculando contagens individualmente para garantir precisão absoluta
    total = Contrato.query.count()
    ativos = Contrato.query.filter(Contrato.status.in_(["Em dia", "Pago"])).count()
    atrasados = Contrato.query.filter(Contrato.status == "Em atraso").count()
    cancelados_inad = Contrato.query.filter(Contrato.status == "Cancelado por Inadimplência").count()
    cancelados_regra = Contrato.query.filter(Contrato.status == "Cancelado por Regra").count()
    # Mortos removido
    
    status_counts = {
        "Ativos": ativos,
        "Atrasados": atrasados,
        "Cancelados_Inad": cancelados_inad,
        "Cancelados_Regra": cancelados_regra
    }
    
    print(f"DEBUG DASHBOARD: {status_counts}")
    
    return render_template(
        'dashboard.html',
        status_counts=status_counts,
        total=total
    )

# ROTAS DE HIGIENIZAÇÃO REMOVIDAS (Anti-pattern "Mortos")

# Rota para listar detalhadamente os contratos cadastrados
@app.route('/contratos')
def listar_contratos():
    busca = request.args.get("busca", "")
    responsavel = request.args.get("responsavel", "")
    status = request.args.get("status", "")
    parcela = request.args.get("parcela", type=int)
    atraso_min = request.args.get("atraso_min", type=int)
    atraso_max = request.args.get("atraso_max", type=int)
    ordenar_por = request.args.get("ordenar", "dias_atraso_desc")
    critico = request.args.get("critico", "")
    sms = request.args.get("sms", "")

    # OTIMIZAÇÃO: Usar outerjoin para não perder contratos sem vendedor
    query = db.session.query(Contrato, Vendedor.nome.label('nome_vendedor')).outerjoin(Vendedor)

    if busca:
        busca_limpa = busca.strip()
        # OTIMIZAÇÃO: Se busca for numérica, usar LIKE com prefixo (mais rápido)
        if busca_limpa.isdigit():
            query = query.filter(
                db.or_(
                    Contrato.proposta.like(f"{busca_limpa}%"),
                    Contrato.contrato.like(f"{busca_limpa}%"),
                    Contrato.cnpj_cpf.like(f"{busca_limpa}%"),
                    Contrato.celular.like(f"{busca_limpa}%")
                )
            )
        else:
            # Busca textual - ILIKE quando necessário
            query = query.filter(
                db.or_(
                    Contrato.proposta.ilike(f"%{busca_limpa}%"),
                    Contrato.contrato.ilike(f"%{busca_limpa}%"),
                    Contrato.razao_social.ilike(f"%{busca_limpa}%"),
                    Contrato.cnpj_cpf.ilike(f"%{busca_limpa}%"),
                    Contrato.celular.ilike(f"%{busca_limpa}%")
                )
            )

    if status == "Em dia + Pago":
        query = query.filter(Contrato.status.in_(["Em dia", "Pago"]))
    elif status == "Cancelado Total":
        query = query.filter(Contrato.status.in_(["Cancelado por Inadimplência", "Cancelado por Regra"]))
    elif status:
        query = query.filter(Contrato.status == status)

    if responsavel:
        subquery_ids = db.session.query(ResponsavelCobranca.contrato_id).filter_by(usuario=responsavel).subquery()
        query = query.filter(Contrato.id.in_(db.session.query(subquery_ids.c.contrato_id)))

    if critico == "1":
        query = query.filter(Contrato.cliente_critico == True)
    elif critico == "0":
        query = query.filter(Contrato.cliente_critico == False)

    if sms == "1":
        query = query.filter(Contrato.envio_sms == True)
    elif sms == "0":
        query = query.filter(Contrato.envio_sms == False)

    if parcela:
        query = query.filter(Contrato.parcela_atual == parcela)

    if atraso_min is not None:
        query = query.filter(Contrato.dias_atraso >= atraso_min)
    if atraso_max is not None:
        query = query.filter(Contrato.dias_atraso <= atraso_max)

    if ordenar_por == "dias_atraso_asc":
        query = query.order_by(Contrato.dias_atraso.asc().nullslast())
    elif ordenar_por == "dias_atraso_desc":
        query = query.order_by(Contrato.dias_atraso.desc().nullslast())
    elif ordenar_por == "parcela_asc":
        query = query.order_by(Contrato.parcela_atual.asc().nullslast())
    elif ordenar_por == "parcela_desc":
        query = query.order_by(Contrato.parcela_atual.desc().nullslast())

    # OTIMIZAÇÃO: Adicionar paginação para melhorar performance
    page = request.args.get('page', 1, type=int)
    per_page = 100  # Ajuste conforme necessário
    resultados_paginados = query.paginate(page=page, per_page=per_page, error_out=False)
    resultados = resultados_paginados.items

    responsaveis = db.session.query(ResponsavelCobranca.usuario).distinct().all()

    return render_template(
        "index.html",
        resultados=resultados,
        busca=busca,
        status=status,
        responsavel=responsavel,
        parcela=parcela,
        atraso_min=atraso_min,
        atraso_max=atraso_max,
        critico=critico,
        sms=sms,
        ordenar_por=ordenar_por,
        responsaveis=[r[0] for r in responsaveis],
        pagination=resultados_paginados
    )


@app.route("/cobranca")
def painel_cobranca():
    session.pop('_flashes', None)

    if request.args:
        session["filtros_cobranca"] = request.args.to_dict()

    filtros = session.get("filtros_cobranca", {})

    busca = filtros.get("busca", "")
    responsavel = filtros.get("responsavel", "")
    status = filtros.get("status", "")
    try:
        parcela = int(filtros.get("parcela")) if filtros.get("parcela") else None
    except ValueError:
        parcela = None

    try:
        atraso_min = int(filtros.get("atraso_min")) if filtros.get("atraso_min") else None
    except ValueError:
        atraso_min = None

    try:
        atraso_max = int(filtros.get("atraso_max")) if filtros.get("atraso_max") else None
    except ValueError:
        atraso_max = None
    ordenar_por = filtros.get("ordenar", "dias_atraso_desc")
    critico = filtros.get("critico", "")
    sms = filtros.get("sms", "")

    # Base query: contratos em atraso (padrão) - Mortos removido
    query = Contrato.query.filter(Contrato.status == "Em atraso")

    # OTIMIZAÇÃO: Busca mais eficiente - usar LIKE com prefixo quando possível
    if busca:
        busca_limpa = busca.strip()
        if busca_limpa.isdigit():
            # Busca numérica - usar LIKE com prefixo (mais rápido, usa índices)
            query = query.filter(
                db.or_(
                    Contrato.proposta.like(f"{busca_limpa}%"),
                    Contrato.contrato.like(f"{busca_limpa}%"),
                    Contrato.cnpj_cpf.like(f"{busca_limpa}%"),
                    Contrato.celular.like(f"{busca_limpa}%")
                )
            )
        else:
            # Busca textual - ILIKE quando necessário
            query = query.filter(
                db.or_(
                    Contrato.proposta.ilike(f"%{busca_limpa}%"),
                    Contrato.contrato.ilike(f"%{busca_limpa}%"),
                    Contrato.razao_social.ilike(f"%{busca_limpa}%"),
                    Contrato.cnpj_cpf.ilike(f"%{busca_limpa}%"),
                    Contrato.celular.ilike(f"%{busca_limpa}%")
                )
            )

    if status:
        query = query.filter(Contrato.status == status)

    if responsavel:
        subquery_ids = db.session.query(ResponsavelCobranca.contrato_id).filter_by(usuario=responsavel).subquery()
        query = query.filter(Contrato.id.in_(db.session.query(subquery_ids.c.contrato_id)))
    
    if critico == "1":
        query = query.filter(Contrato.cliente_critico == True)
    elif critico == "0":
        query = query.filter(Contrato.cliente_critico == False)

    if sms == "1":
        query = query.filter(Contrato.envio_sms == True)
    elif sms == "0":
        query = query.filter(Contrato.envio_sms == False)

    if parcela:
        query = query.filter(Contrato.parcela_atual == parcela)

    if atraso_min is not None:
        query = query.filter(Contrato.dias_atraso >= atraso_min)
    if atraso_max is not None:
        query = query.filter(Contrato.dias_atraso <= atraso_max)

    # Ordenação com nullslast para melhor performance
    if ordenar_por == "dias_atraso_asc":
        query = query.order_by(Contrato.dias_atraso.asc().nullslast())
    elif ordenar_por == "dias_atraso_desc":
        query = query.order_by(Contrato.dias_atraso.desc().nullslast())
    elif ordenar_por == "parcela_asc":
        query = query.order_by(Contrato.parcela_atual.asc().nullslast())
    elif ordenar_por == "parcela_desc":
        query = query.order_by(Contrato.parcela_atual.desc().nullslast())

    # OTIMIZAÇÃO CRÍTICA: Adicionar paginação
    page = request.args.get('page', 1, type=int)
    per_page = 100  # Ajuste conforme necessário (50-200)
    contratos_paginados = query.paginate(page=page, per_page=per_page, error_out=False)
    contratos = contratos_paginados.items
    
    # OTIMIZAÇÃO: Carregar todas as ações e responsáveis de uma vez (evita N+1 queries)
    if contratos:
        contratos_ids = [c.id for c in contratos]
        
        # OTIMIZAÇÃO: Usar DISTINCT ON (PostgreSQL) em vez de window function - mais rápido
        # Alternativa mais eficiente: buscar apenas os IDs das últimas ações
        from sqlalchemy import func
        
        # Subquery para encontrar o ID da última ação de cada contrato
        subquery_max_dia = db.session.query(
            AcaoCobranca.contrato_id,
            func.max(AcaoCobranca.dia_atraso).label('max_dia')
        ).filter(
            AcaoCobranca.contrato_id.in_(contratos_ids)
        ).group_by(AcaoCobranca.contrato_id).subquery()
        
        # Buscar as ações correspondentes
        ultimas_acoes_query = db.session.query(AcaoCobranca).join(
            subquery_max_dia,
            (AcaoCobranca.contrato_id == subquery_max_dia.c.contrato_id) &
            (AcaoCobranca.dia_atraso == subquery_max_dia.c.max_dia)
        ).all()
        
        # Criar dicionário para acesso rápido
        ultimas_acoes_dict = {}
        for acao in ultimas_acoes_query:
            # Se houver múltiplas ações com mesmo dia_atraso, pegar a mais recente
            if acao.contrato_id not in ultimas_acoes_dict:
                ultimas_acoes_dict[acao.contrato_id] = acao
            elif acao.enviada_em and ultimas_acoes_dict[acao.contrato_id].enviada_em:
                if acao.enviada_em > ultimas_acoes_dict[acao.contrato_id].enviada_em:
                    ultimas_acoes_dict[acao.contrato_id] = acao
        
        # Buscar todos os responsáveis de uma vez
        responsaveis_query = ResponsavelCobranca.query.filter(
            ResponsavelCobranca.contrato_id.in_(contratos_ids)
        ).all()
        responsaveis_dict = {r.contrato_id: r.usuario for r in responsaveis_query}
    else:
        ultimas_acoes_dict = {}
        responsaveis_dict = {}
    
    # OTIMIZAÇÃO: Cachear lista de responsáveis (não muda frequentemente)
    responsaveis = db.session.query(ResponsavelCobranca.usuario).distinct().all()
    
    # Montar lista sem queries adicionais
    lista = []
    for c in contratos:
        lista.append({
            "contrato": c,
            "ultima_acao": ultimas_acoes_dict.get(c.id),
            "responsavel": responsaveis_dict.get(c.id)
        })

    return render_template(
        "cobranca.html",
        contratos=lista,
        responsaveis=[r[0] for r in responsaveis],
        busca=busca,
        status=status,
        responsavel=responsavel,
        parcela=parcela,
        atraso_min=atraso_min,
        atraso_max=atraso_max,
        critico=critico,
        sms=sms,
        ordenar_por=ordenar_por,
        pagination=contratos_paginados  # Adicionar paginação ao template
    )

@app.route("/toggle_sms/<int:contrato_id>")
def toggle_sms(contrato_id):
    contrato = Contrato.query.get_or_404(contrato_id)
    contrato.envio_sms = not contrato.envio_sms
    db.session.commit()
    flash(f"Envio de SMS {'ativado' if contrato.envio_sms else 'desativado'} para este contrato.")
    return redirect(f"/historico/{contrato_id}")

@app.route("/toggle_critico/<int:contrato_id>")
def toggle_critico(contrato_id):
    contrato = Contrato.query.get_or_404(contrato_id)
    contrato.cliente_critico = not contrato.cliente_critico
    db.session.commit()
    flash(f"Cliente {'marcado como CRÍTICO' if contrato.cliente_critico else 'removido de críticos'}.")
    return redirect(f"/historico/{contrato_id}")

@app.route("/historico/<int:contrato_id>")
def historico_cobranca(contrato_id):
    contrato = Contrato.query.get_or_404(contrato_id)
    acoes = AcaoCobranca.query.filter_by(contrato_id=contrato_id).order_by(AcaoCobranca.dia_atraso).all()
    return render_template("historico.html", contrato=contrato, acoes=acoes)

from flask import request, jsonify

@app.route("/assumir/<int:contrato_id>", methods=["GET", "POST"])
def assumir_contrato(contrato_id):
    contrato = Contrato.query.get_or_404(contrato_id)
    usuario_logado = session.get("usuario_nome", "Usuário de Cobrança")

    # Verifica se já existe responsável
    responsavel_existente = ResponsavelCobranca.query.filter_by(contrato_id=contrato.id).first()

    if responsavel_existente:
        responsavel_existente.usuario = usuario_logado
    else:
        novo = ResponsavelCobranca(contrato_id=contrato.id, usuario=usuario_logado)
        db.session.add(novo)

    db.session.commit()

    # Se for requisição via fetch(), retorna JSON
    if request.headers.get("X-Requested-With") == "XMLHttpRequest":
        return jsonify({"status": "ok", "responsavel": usuario_logado})

    # Se for acesso direto, redireciona normalmente
    flash(f"Contrato {contrato.contrato} agora está sob responsabilidade de {usuario_logado}")
    return redirect("/cobranca")

@app.route("/registrar_acao/<int:contrato_id>", methods=["GET", "POST"])
def registrar_acao(contrato_id):
    # OTIMIZAÇÃO: Carregar vendedor junto com contrato
    contrato = Contrato.query.options(joinedload(Contrato.vendedor)).get_or_404(contrato_id)

    # OTIMIZAÇÃO: Usar relacionamento em vez de query separada
    vendedor = contrato.vendedor.nome if contrato.vendedor else "-"

    if request.method == "POST":
        tipo = request.form.get("tipo")
        mensagem = request.form.get("mensagem")
        usuario_logado = session.get("usuario_nome", "Usuário de Cobrança")

        if not mensagem:
            flash("A mensagem é obrigatória.")
            return redirect(f"/registrar_acao/{contrato_id}")

        nova_acao = AcaoCobranca(
            contrato_id=contrato.id,
            tipo=tipo,
            mensagem=mensagem,
            dia_atraso=contrato.dias_atraso or 0,
            parcela=contrato.parcela_atual,
            enviada_em=agora_brasil(),
            status_envio="Manual",
            usuario=usuario_logado
        )
        db.session.add(nova_acao)
        db.session.commit()

        flash("Ação registrada com sucesso.")
        anchor = request.args.get("anchor")
        return redirect(f"/cobranca#{anchor}" if anchor else "/cobranca")

    return render_template("registrar_acao.html", contrato=contrato, vendedor=vendedor)

@app.route("/editar_contrato/<int:contrato_id>", methods=["GET", "POST"])
@login_required
@admin_required
def editar_contrato(contrato_id):
    contrato = Contrato.query.get_or_404(contrato_id)
    vendedores = Vendedor.query.order_by(Vendedor.nome).all()

    if request.method == "POST":
        try:
            contrato.celular = request.form.get("celular")
            contrato.email = request.form.get("email")
            contrato.razao_social = request.form.get("razao_social")
            contrato.nome_plano = request.form.get("nome_plano")
            
            vidas_str = request.form.get("vidas")
            if vidas_str: contrato.vidas = int(vidas_str)
            
            valor_str = request.form.get("valor_parcela", "").replace("R$","").replace(".","").replace(",",".")
            if valor_str: contrato.valor_parcela = float(valor_str)
            
            parcela_str = request.form.get("parcela_atual")
            if parcela_str: contrato.parcela_atual = int(parcela_str)
            
            contrato.status = request.form.get("status")
            
            vendedor_id = request.form.get("vendedor_id")
            if vendedor_id: contrato.vendedor_id = int(vendedor_id)

            db.session.commit()
            flash("Contrato atualizado com sucesso!", "success")
            return redirect(url_for("cobranca"))
            
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar: {e}", "error")

    return render_template("editar_contrato.html", contrato=contrato, vendedores=vendedores)


@app.route("/relatorio_cobranca")
@login_required
@admin_required
def relatorio_cobranca():
    from datetime import datetime
    from sqlalchemy import extract, func, case

    mes = request.args.get("mes", datetime.today().month, type=int)
    ano = request.args.get("ano", datetime.today().year, type=int)
    usuario_filtro = request.args.get("usuario", "")

    hoje = datetime.today()
    primeiro_dia = datetime(ano, mes, 1)
    if mes == 12:
        proximo_mes = datetime(ano + 1, 1, 1)
    else:
        proximo_mes = datetime(ano, mes + 1, 1)

    # Contratos com ações manuais registradas no mês
    contratos_em_atraso = db.session.query(AcaoCobranca.contrato_id).filter(
        AcaoCobranca.enviada_em >= primeiro_dia,
        AcaoCobranca.enviada_em < proximo_mes
    ).distinct().all()
    contratos_ids = {c[0] for c in contratos_em_atraso}

    # Adicionar contratos cancelados no mês
    contratos_cancelados = Contrato.query.filter(
        Contrato.mes_cancelamento >= primeiro_dia,
        Contrato.mes_cancelamento < proximo_mes,
        Contrato.status.ilike('%Cancelado%')
    ).all()

    for contrato in contratos_cancelados:
        contratos_ids.add(contrato.id)

    contratos = Contrato.query.filter(Contrato.id.in_(contratos_ids)).all()

    # OTIMIZAÇÃO: Métricas calculadas no banco em vez de Python
    if contratos_ids:
        stats = db.session.query(
            func.count(Contrato.id).label('total'),
            func.sum(case((Contrato.status.in_(["Pago", "Em dia"]), 1), else_=0)).label('pagos'),
            func.sum(case((Contrato.status.ilike('%Cancelado%'), 1), else_=0)).label('cancelados'),
            func.sum(case((Contrato.status == "Em atraso", 1), else_=0)).label('em_atraso')
        ).filter(Contrato.id.in_(contratos_ids)).first()
        
        pagos = stats.pagos or 0
        cancelados = stats.cancelados or 0
        ainda_em_atraso = stats.em_atraso or 0
        total = stats.total or 0
    else:
        pagos = cancelados = ainda_em_atraso = total = 0

    taxa_recuperacao = f"{round((pagos / total * 100), 2)}%" if total > 0 else "0%"
    taxa_cancelamento = f"{round((cancelados / total * 100), 2)}%" if total > 0 else "0%"

    acoes_no_mes = AcaoCobranca.query.filter(
        AcaoCobranca.enviada_em >= primeiro_dia,
        AcaoCobranca.enviada_em < proximo_mes
    ).count()

    usuarios = db.session.query(ResponsavelCobranca.usuario).distinct().all()

    if usuario_filtro:
        usuarios = [(usuario_filtro,)]

    resumo_usuarios = []

    # OTIMIZAÇÃO: Carregar todos os contratos de uma vez e fazer cálculos em batch
    todos_contratos_ids = []
    for (usuario,) in usuarios:
        contratos_ids_subquery = db.session.query(ResponsavelCobranca.contrato_id).filter_by(usuario=usuario).subquery()
        contratos_ids_usuario = [c[0] for c in db.session.query(contratos_ids_subquery.c.contrato_id).all()]
        todos_contratos_ids.extend(contratos_ids_usuario)
    
    # Carregar todos os contratos de uma vez
    if todos_contratos_ids:
        todos_contratos_dict = {c.id: c for c in Contrato.query.filter(Contrato.id.in_(todos_contratos_ids)).all()}
    else:
        todos_contratos_dict = {}
    
    # OTIMIZAÇÃO: Buscar todas as ações e últimas ações de uma vez
    acoes_por_usuario = {}
    ultimas_acoes_por_usuario = {}
    
    if usuarios:
        usuarios_list = [u[0] for u in usuarios]
        acoes_todas = AcaoCobranca.query.filter(
            AcaoCobranca.usuario.in_(usuarios_list),
            extract("month", AcaoCobranca.enviada_em) == mes,
            extract("year", AcaoCobranca.enviada_em) == ano
        ).all()
        
        # Agrupar ações por usuário
        for acao in acoes_todas:
            if acao.usuario not in acoes_por_usuario:
                acoes_por_usuario[acao.usuario] = []
            acoes_por_usuario[acao.usuario].append(acao)
        
        # Buscar últimas ações de uma vez
        ultimas_acoes_query = db.session.query(
            AcaoCobranca.usuario,
            func.max(AcaoCobranca.enviada_em).label('ultima_data')
        ).filter(
            AcaoCobranca.usuario.in_(usuarios_list),
            extract("month", AcaoCobranca.enviada_em) == mes,
            extract("year", AcaoCobranca.enviada_em) == ano
        ).group_by(AcaoCobranca.usuario).all()
        
        ultimas_acoes_por_usuario = {u[0]: u[1] for u in ultimas_acoes_query}

    for (usuario,) in usuarios:
        contratos_ids_subquery = db.session.query(ResponsavelCobranca.contrato_id).filter_by(usuario=usuario).subquery()
        contratos_ids_usuario = [c[0] for c in db.session.query(contratos_ids_subquery.c.contrato_id).all()]
        
        # Usar contratos já carregados
        contratos_usuario = [todos_contratos_dict.get(cid) for cid in contratos_ids_usuario if todos_contratos_dict.get(cid)]
        
        # OTIMIZAÇÃO: Calcular métricas usando os contratos já carregados
        pagos_usuario = sum(1 for c in contratos_usuario if c and c.status in ["Pago", "Em dia"])
        em_atraso_usuario = sum(1 for c in contratos_usuario if c and c.status == "Em atraso")
        cancelados_usuario = sum(1 for c in contratos_usuario if c and c.status and "Cancelado" in c.status)
        total_assumidos = len(contratos_usuario)

        acoes = acoes_por_usuario.get(usuario, [])
        ultima_acao_data = ultimas_acoes_por_usuario.get(usuario)
        ultima_acao = ultima_acao_data.strftime('%d/%m/%Y') if ultima_acao_data else None

        resumo_usuarios.append({
            "usuario": usuario,
            "assumidos": total_assumidos,
            "acoes": len(acoes),
            "ultima_acao": ultima_acao,
            "pagos": pagos_usuario,
            "em_atraso": em_atraso_usuario,
            "cancelados": cancelados_usuario
        })

    return render_template(
        "relatorio_cobranca.html",
        total=total,
        pagos=pagos,
        cancelados=cancelados,
        ainda_em_atraso=ainda_em_atraso,
        taxa_recuperacao=taxa_recuperacao,
        taxa_cancelamento=taxa_cancelamento,
        acoes=acoes_no_mes,
        resumo_usuarios=resumo_usuarios,
        mes=mes,
        ano=ano,
        usuarios_filtro=[u[0] for u in usuarios],
        usuario_selecionado=usuario_filtro
    )

@app.route("/exportar_relatorio")
def exportar_relatorio():
    from datetime import datetime
    from sqlalchemy import extract

    mes = request.args.get("mes", datetime.today().month, type=int)
    ano = request.args.get("ano", datetime.today().year, type=int)
    usuario = request.args.get("usuario", "")

    subquery = db.session.query(
        AcaoCobranca.contrato_id,
        db.func.max(AcaoCobranca.enviada_em).label("ultima_data")
    ).group_by(AcaoCobranca.contrato_id).subquery()

    acoes = db.session.query(AcaoCobranca).join(
        subquery,
        (AcaoCobranca.contrato_id == subquery.c.contrato_id) &
        (AcaoCobranca.enviada_em == subquery.c.ultima_data)
    ).filter(
        extract("month", AcaoCobranca.enviada_em) == mes,
        extract("year", AcaoCobranca.enviada_em) == ano
    )

    if usuario:
        acoes = acoes.filter(AcaoCobranca.usuario == usuario)

    acoes = acoes.all()
    hoje = datetime.today().date()

    # OTIMIZAÇÃO: Carregar todos os contratos de uma vez com eager loading
    if acoes:
        contratos_ids = [acao.contrato_id for acao in acoes]
        contratos_dict = {
            c.id: c for c in Contrato.query.options(joinedload(Contrato.vendedor)).filter(
                Contrato.id.in_(contratos_ids)
            ).all()
        }
    else:
        contratos_dict = {}

    dados = []
    for acao in acoes:
        contrato = contratos_dict.get(acao.contrato_id)
        if not contrato:
            continue
        dados.append({
            "Contrato": contrato.contrato,
            "Proposta": contrato.proposta,
            "Razão Social": contrato.razao_social,
            "CNPJ/CPF": contrato.cnpj_cpf,
            "Cliente Crítico": "Sim" if contrato.cliente_critico else "Não",
            "Atividade Econômica": contrato.atividade_economica or "",
            "Celular": contrato.celular,
            "E-mail": contrato.email,
            "Cidade": contrato.cidade,
            "Data de Vigência": contrato.data_vigencia.strftime("%d/%m/%Y") if contrato.data_vigencia else "",
            "Vidas": contrato.vidas,
            "Valor da Parcela": contrato.valor_parcela,
            "Parcela Atual": contrato.parcela_atual,
            "Dias de Atraso": contrato.dias_atraso,
            "Status Atual": contrato.status,
            "Mês de Cancelamento": contrato.mes_cancelamento.strftime("%m/%Y") if contrato.mes_cancelamento else "",
            "Vendedor": contrato.vendedor.nome if contrato.vendedor else "",
            "Última Ação": acao.tipo,
            "Responsável": acao.usuario or "-",
            "Data da Ação": acao.enviada_em.strftime("%d/%m/%Y %H:%M"),
            "Mensagem Enviada": acao.mensagem
        })

    df = pd.DataFrame(dados)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Relatório")

    output.seek(0)
    return send_file(
        output,
        download_name=f"relatorio_cobranca_{hoje.strftime('%Y_%m_%d')}.xlsx",
        as_attachment=True
    )


@app.route("/exportar_relatorio_atendentes")
def exportar_relatorio_atendentes():
    from datetime import datetime
    from sqlalchemy import extract, func, case

    hoje = datetime.today().date()
    mes = request.args.get("mes", hoje.month, type=int)
    ano = request.args.get("ano", hoje.year, type=int)
    usuario_filtro = request.args.get("usuario", "")

    usuarios = db.session.query(ResponsavelCobranca.usuario).distinct().all()
    if usuario_filtro:
        usuarios = [(usuario_filtro,)]

    # OTIMIZAÇÃO: Carregar todos os contratos de uma vez
    todos_contratos_ids = []
    usuarios_list = [u[0] for u in usuarios]
    
    for (usuario,) in usuarios:
        contratos_ids = db.session.query(ResponsavelCobranca.contrato_id).filter_by(usuario=usuario).all()
        todos_contratos_ids.extend([c[0] for c in contratos_ids])
    
    # Carregar todos os contratos de uma vez
    if todos_contratos_ids:
        todos_contratos_dict = {c.id: c for c in Contrato.query.filter(Contrato.id.in_(todos_contratos_ids)).all()}
    else:
        todos_contratos_dict = {}
    
    # OTIMIZAÇÃO: Buscar contagem de ações de uma vez
    acoes_por_usuario = {}
    if usuarios_list:
        acoes_query = db.session.query(
            AcaoCobranca.usuario,
            func.count(AcaoCobranca.id).label('total')
        ).filter(
            AcaoCobranca.usuario.in_(usuarios_list),
            extract("month", AcaoCobranca.enviada_em) == mes,
            extract("year", AcaoCobranca.enviada_em) == ano
        ).group_by(AcaoCobranca.usuario).all()
        
        acoes_por_usuario = {u[0]: u[1] for u in acoes_query}

    linhas = []
    for (usuario,) in usuarios:
        contratos_ids = [c[0] for c in db.session.query(ResponsavelCobranca.contrato_id).filter_by(usuario=usuario).all()]
        contratos = [todos_contratos_dict.get(cid) for cid in contratos_ids if todos_contratos_dict.get(cid)]

        total_assumidos = len(contratos)
        pagos = sum(1 for c in contratos if c and c.status in ["Pago", "Em dia"])
        em_atraso = sum(1 for c in contratos if c and c.status == "Em atraso")
        cancelados = sum(1 for c in contratos if c and c.status and "Cancelado" in c.status)

        acoes = acoes_por_usuario.get(usuario, 0)

        taxa = f"{round((pagos / total_assumidos * 100), 2)}%" if total_assumidos > 0 else "0%"

        linhas.append({
            "Atendente": usuario,
            "Contratos Assumidos": total_assumidos,
            "Ações Realizadas": acoes,
            "Contratos Pagos": pagos,
            "Contratos Ainda em Atraso": em_atraso,
            "Contratos Cancelados": cancelados,
            "Taxa de Recuperação": taxa
        })

    df = pd.DataFrame(linhas)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Atendentes")

    output.seek(0)
    return send_file(
        output,
        download_name=f"relatorio_atendentes_{hoje.strftime('%Y_%m_%d')}.xlsx",
        as_attachment=True
    )

@app.route("/editar_contrato", methods=["GET", "POST"])
def editar_contrato_busca():
    if request.method == "POST":
        contrato_numero = request.form.get("numero_contrato")
        contrato = Contrato.query.filter_by(contrato=contrato_numero).first()
        if contrato:
            return redirect(f"/editar_contrato/{contrato.id}")
        else:
            flash("Contrato não encontrado.")
            return redirect("/editar_contrato")

    return render_template("buscar_contrato.html")

@app.route("/editar_contrato/<int:contrato_id>", methods=["GET", "POST"])
def editar_cadastro_contrato(contrato_id):
    contrato = Contrato.query.get_or_404(contrato_id)

    if request.method == "POST":
        usuario_tipo = session.get("usuario_tipo", "")

        contrato.celular = request.form.get("celular")
        contrato.email = request.form.get("email")

        if usuario_tipo == "admin":
            contrato.razao_social = request.form.get("razao_social")
            contrato.nome_plano = request.form.get("nome_plano")
            contrato.vidas = request.form.get("vidas", type=int)
            contrato.valor_parcela = request.form.get("valor_parcela", type=float)
            contrato.parcela_atual = request.form.get("parcela_atual", type=int)
            contrato.status = request.form.get("status")
            id_vendedor = request.form.get("vendedor_id")
            if id_vendedor:
                contrato.vendedor_id = int(id_vendedor)

        db.session.commit()
        flash("Contrato atualizado com sucesso.")
        return redirect("/")

    vendedores = Vendedor.query.all()
    return render_template("editar_contrato.html", contrato=contrato, vendedores=vendedores)

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email")
        senha = request.form.get("senha")

        usuario = Usuario.query.filter_by(email=email).first()
        if usuario and check_password_hash(usuario.senha, senha):
            session["usuario_id"] = usuario.id
            session["usuario_nome"] = usuario.nome
            session["usuario_tipo"] = usuario.tipo
            flash("Login realizado com sucesso.")
            return redirect("/")
        else:
            flash("E-mail ou senha inválidos.")

    return render_template("login.html")

@app.route("/cadastro_usuario", methods=["GET", "POST"])
@login_required
@admin_required
def cadastro_usuario():
    if request.method == "POST":
        nome = request.form.get("nome")
        email = request.form.get("email")
        senha = request.form.get("senha")
        tipo = request.form.get("tipo")

        if not nome or not email or not senha or tipo not in ["admin", "usuario"]:
            flash("Todos os campos são obrigatórios e o tipo deve ser válido.")
            return redirect("/cadastro_usuario")

        existente = Usuario.query.filter_by(email=email).first()
        if existente:
            flash("Já existe um usuário com este e-mail.")
            return redirect("/cadastro_usuario")

        novo_usuario = Usuario(
            nome=nome,
            email=email,
            senha=generate_password_hash(senha),
            tipo=tipo
        )
        db.session.add(novo_usuario)
        db.session.commit()
        flash("Usuário cadastrado com sucesso!")
        return redirect("/cadastro_usuario")

    return render_template("cadastro_usuario.html")

@app.route("/relatorio_vivos")
def relatorio_vivos():
    if session.get("usuario_tipo") != "admin":
        flash("Acesso restrito ao administrador.")
        return redirect("/")
    # OTIMIZAÇÃO: Carregar vendedores junto com contratos (eager loading)
    contratos_vivos = Contrato.query.options(joinedload(Contrato.vendedor)).filter(
        Contrato.status.in_(["Em dia", "Pago", "Em atraso"])
    ).all()
    
    total_contratos = len(contratos_vivos)
    total_vidas = sum(c.vidas or 0 for c in contratos_vivos)
    total_faturamento = sum(c.valor_parcela or Decimal(0) for c in contratos_vivos)
    
    def calcular_comissao(contrato):
        if contrato.parcela_atual in [1, 2, 3]:
            return contrato.valor_parcela or Decimal(0)  # agenciamento = 100%
        elif contrato.parcela_atual and contrato.parcela_atual >= 4:
            return (contrato.valor_parcela or Decimal(0)) * Decimal("0.04")  # vitalício = 4%
        return 0

    total_agenciamento = sum(c.valor_parcela or Decimal(0) for c in contratos_vivos if c.parcela_atual in [1, 2, 3])
    total_vitalicio = sum((c.valor_parcela or Decimal(0)) * Decimal("0.04") for c in contratos_vivos if c.parcela_atual and c.parcela_atual >= 4)

    # agrupamentos
    from collections import defaultdict, Counter

    vidas_por_produto = defaultdict(int)
    agenciamento_por_produto = defaultdict(lambda: Decimal(0))
    vitalicio_por_produto = defaultdict(lambda: Decimal(0))
    vidas_por_vendedora = defaultdict(int)
    contratos_por_vendedora = defaultdict(int)
    atividades = Counter()
    cidades = Counter()
    vidas_por_contrato_total = []
    ticket_por_contrato = []
    ticket_por_vida = []
    comissao_total = []

    for c in contratos_vivos:
        nome_plano = c.nome_plano.strip() if c.nome_plano and c.nome_plano.strip().lower() != "nan" else "Não informado"
        vidas_por_produto[nome_plano] += c.vidas or 0
        if c.parcela_atual in [1, 2, 3]:
            agenciamento_por_produto[nome_plano] += c.valor_parcela or Decimal(0)
        elif c.parcela_atual and c.parcela_atual >= 4:
            vitalicio_por_produto[nome_plano] += (c.valor_parcela or Decimal(0)) * Decimal("0.04")
        
        # Vendedor já está carregado, não precisa query adicional
        if c.vendedor:
            vidas_por_vendedora[c.vendedor.nome] += c.vidas or 0
            contratos_por_vendedora[c.vendedor.nome] += 1

        atividades[c.atividade_economica or "-"] += 1
        cidades[c.cidade or "-"] += 1

        vidas_por_contrato_total.append(c.vidas or 0)
        ticket_por_contrato.append(c.valor_parcela or Decimal(0))
        if c.vidas:
            ticket_por_vida.append((c.valor_parcela or Decimal(0)) / c.vidas)
            comissao_total.append(calcular_comissao(c))

    from statistics import mean, mode

    try:
        moda_vidas = mode(vidas_por_contrato_total)
    except:
        moda_vidas = "Sem valor único"

    percentual_por_produto = {}
    for produto, vidas in vidas_por_produto.items():
        percentual = (vidas / total_vidas * 100) if total_vidas else 0
        percentual_por_produto[produto] = round(percentual, 1)

    total_vidas_com_comissao = sum(c.vidas for c in contratos_vivos if c.vidas)
    comissao_total_somada = sum(comissao_total)
    comissao_media_vida = round(comissao_total_somada / total_vidas_com_comissao, 2) if total_vidas_com_comissao else 0

    media_vidas_por_contrato = round(total_vidas / total_contratos, 2) if total_contratos else 0

    parcela_media = round(mean([c.valor_parcela for c in contratos_vivos if c.valor_parcela]), 2) if contratos_vivos else 0

    parcelas_ativas = [c.parcela_atual for c in contratos_vivos if c.parcela_atual]

    parcela_mediana = round(median(parcelas_ativas), 2) if parcelas_ativas else "-"

    parcela_maxima = max(parcelas_ativas) if parcelas_ativas else "-"

    return render_template("relatorio_vivos.html",
        total_contratos=total_contratos,
        total_vidas=total_vidas,
        media_vidas_por_contrato=media_vidas_por_contrato,
        total_faturamento=round(total_faturamento, 2),
        total_agenciamento=round(total_agenciamento, 2),
        total_vitalicio=round(total_vitalicio, 2),
        vidas_por_produto=dict(vidas_por_produto),
        parcela_mediana=parcela_mediana,
        parcela_maxima=parcela_maxima,
        percentual_por_produto=percentual_por_produto,
        agenciamento_por_produto=dict(agenciamento_por_produto),
        vitalicio_por_produto=dict(vitalicio_por_produto),
        vidas_por_vendedora=dict(vidas_por_vendedora),
        contratos_por_vendedora=dict(contratos_por_vendedora),
        vidas_por_contrato_total=vidas_por_contrato_total,
        parcela_media=parcela_media,
        parcela_moda=moda_vidas,
        ticket_medio_contrato=round(mean(ticket_por_contrato), 2) if ticket_por_contrato else 0,
        ticket_medio_vida=round(mean(ticket_por_vida), 2) if ticket_por_vida else 0,
        comissao_media_contrato=round(mean(comissao_total), 2) if comissao_total else 0,
        comissao_media_vida=comissao_media_vida,
        comissao_agenciamento=round(total_agenciamento, 2),
        comissao_vitalicio=round(total_vitalicio, 2),
        top_atividades=atividades.most_common(10),
        top_cidades=cidades.most_common(10)
    )

@app.route("/logout")
def logout():
    session.clear()
    flash("Você saiu do sistema com sucesso.")
    return redirect(url_for("login"))

@app.route("/limpar_filtros_cobranca")
def limpar_filtros_cobranca():
    session.pop("filtros_cobranca", None)
    return redirect("/cobranca")

@app.route("/exportar_base_completa")
@login_required
@admin_required
def exportar_base_completa():
    """Exporta a base de dados filtrada em um arquivo Excel formatado"""
    from datetime import datetime
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        hoje = datetime.today().date()
        
        # Recupera filtro de status da URL
        filtro_status = request.args.get("status", "TODOS")
        
        # Query de contratos com filtro
        query = Contrato.query.options(joinedload(Contrato.vendedor))
        
        if filtro_status and filtro_status != "TODOS":
            lista_status = [s.strip() for s in filtro_status.split(',')]
            query = query.filter(Contrato.status.in_(lista_status))
        
        # Ordenar por status (ordem lógica)
        ordem_status = ["Em dia", "Pago", "Em atraso", "Cancelado por Inadimplência", "Cancelado por Regra"]
        contratos = query.all()
        
        # Ordenar manualmente
        def sort_key(c):
            try:
                return ordem_status.index(c.status) if c.status in ordem_status else 999
            except:
                return 999
        contratos.sort(key=sort_key)
        
        # IDs dos contratos filtrados (para filtrar outras abas)
        ids_contratos = {c.id for c in contratos}
        
        # Criar buffer
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # ==================== ABA CONTRATOS ====================
            dados_contratos = []
            for c in contratos:
                try:
                    dados_contratos.append({
                        "ID": c.id,
                        "Proposta": c.proposta or "",
                        "Contrato": c.contrato or "",
                        "Data Checagem": c.data_checagem.strftime("%d/%m/%Y") if c.data_checagem else "",
                        "Razão Social": c.razao_social or "",
                        "CNPJ/CPF": c.cnpj_cpf or "",
                        "Celular": c.celular or "",
                        "E-mail": c.email or "",
                        "Atividade Econômica": c.atividade_economica or "",
                        "Cidade": c.cidade or "",
                        "Nome do Plano": c.nome_plano or "",
                        "Data de Vigência": c.data_vigencia.strftime("%d/%m/%Y") if c.data_vigencia else "",
                        "Vidas": c.vidas if c.vidas is not None else "",
                        "Valor da Parcela": float(c.valor_parcela) if c.valor_parcela is not None else "",
                        "Parcela Atual": c.parcela_atual if c.parcela_atual is not None else "",
                        "Status": c.status or "",
                        "Mês de Cancelamento": c.mes_cancelamento.strftime("%m/%Y") if c.mes_cancelamento else "",
                        "Verificado": "Sim" if c.verificado else "Não",
                        "Dias de Atraso": c.dias_atraso if c.dias_atraso is not None else "",
                        "Vendedor": c.vendedor.nome if c.vendedor else "",
                        "Envio SMS": "Sim" if c.envio_sms else "Não",
                        "Cliente Crítico": "Sim" if c.cliente_critico else "Não"
                    })
                except Exception as e:
                    logging.error(f"Erro contrato {c.id}: {e}")
                    continue
            
            if dados_contratos:
                df = pd.DataFrame(dados_contratos)
                df.to_excel(writer, index=False, sheet_name="Contratos")
                
                # Formatação
                ws = writer.sheets["Contratos"]
                header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                thin_border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                
                # Larguras de coluna (ordem: ID, Proposta, Contrato, Data Checagem, Razão Social, CNPJ/CPF, Celular, E-mail, Atividade, Cidade, Plano, Vigência, Vidas, Valor, Parcela, Status, Mês Cancel, Verificado, Dias Atraso, Vendedor, SMS, Crítico)
                col_widths = [8, 12, 12, 14, 45, 18, 15, 30, 25, 18, 25, 14, 8, 14, 10, 28, 14, 12, 12, 25, 10, 12]
                for i, width in enumerate(col_widths, 1):
                    if i <= len(col_widths):
                        ws.column_dimensions[get_column_letter(i)].width = width
                
                # Formatar header
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border
                
                # Formatar dados
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="left")
            
            # ==================== ABA AÇÕES (APENAS DOS CONTRATOS FILTRADOS) ====================
            if ids_contratos:
                acoes = AcaoCobranca.query.filter(AcaoCobranca.contrato_id.in_(ids_contratos)).all()
                if acoes:
                    dados_acoes = []
                    for a in acoes:
                        dados_acoes.append({
                            "Contrato ID": a.contrato_id,
                            "Tipo": a.tipo or "",
                            "Mensagem": (a.mensagem or "")[:100],  # Limitar tamanho
                            "Dia Atraso": a.dia_atraso or 0,
                            "Parcela": a.parcela or 0,
                            "Enviada": a.enviada_em.strftime("%d/%m/%Y") if a.enviada_em else "",
                            "Status": a.status_envio or "",
                        })
                    
                    df_acoes = pd.DataFrame(dados_acoes)
                    df_acoes.to_excel(writer, index=False, sheet_name="Ações")
                    
                    # Formatação
                    ws = writer.sheets["Ações"]
                    for i, width in enumerate([12, 15, 50, 10, 10, 12, 15], 1):
                        ws.column_dimensions[get_column_letter(i)].width = width
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
            
            # ==================== ABA RESPONSÁVEIS (APENAS DOS CONTRATOS FILTRADOS) ====================
            if ids_contratos:
                responsaveis = ResponsavelCobranca.query.filter(
                    ResponsavelCobranca.contrato_id.in_(ids_contratos)
                ).all()
                if responsaveis:
                    dados_resp = [{"Contrato ID": r.contrato_id, "Usuário": r.usuario or ""} for r in responsaveis]
                    df_resp = pd.DataFrame(dados_resp)
                    df_resp.to_excel(writer, index=False, sheet_name="Responsáveis")
                    
                    ws = writer.sheets["Responsáveis"]
                    ws.column_dimensions["A"].width = 12
                    ws.column_dimensions["B"].width = 25
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
            
            # Aba 5: Usuários (sem senhas)
            usuarios = Usuario.query.all()
            dados_usuarios = []
            for u in usuarios:
                dados_usuarios.append({
                    "ID": u.id,
                    "Nome": u.nome or "",
                    "E-mail": u.email or "",
                    "Tipo": u.tipo or ""
                })
            if dados_usuarios:
                df_usuarios = pd.DataFrame(dados_usuarios)
                df_usuarios.to_excel(writer, index=False, sheet_name="Usuários")
        
        output.seek(0)
        return send_file(
            output,
            download_name=f"base_completa_{hoje.strftime('%Y_%m_%d')}.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logging.error(f"Erro ao exportar base completa: {e}", exc_info=True)
        flash(f"Erro ao exportar base: {str(e)}")
        return redirect(url_for("dashboard"))

@app.route("/importar_contratos", methods=["GET", "POST"])
@login_required
@admin_required
def importar_contratos_view():
    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        if not arquivo:
            flash("Selecione um arquivo Excel para importar.")
            return redirect(url_for("importar_contratos_view"))

        if not arquivo.filename.lower().endswith((".xls", ".xlsx")):
            flash("Envie um arquivo Excel (.xls ou .xlsx).")
            return redirect(url_for("importar_contratos_view"))

        try:
            resumo, mensagens = importar_contratos_de_planilha(arquivo)
            flash(resumo)
            for msg in mensagens[:10]:
                flash(msg)
        except Exception as e:
            logging.error(f"Erro ao importar contratos via upload: {e}", exc_info=True)
            flash(f"Erro ao importar contratos: {e}")

        return redirect(url_for("importar_contratos_view"))

    return render_template("importar_contratos.html")


@app.route("/sobrepor_status", methods=["GET", "POST"])
@login_required
@admin_required
def sobrepor_status_view():
    """
    Página para sobrepor status de contratos via planilha.
    
    Permite que o admin force a mudança de status de contratos em massa,
    útil para "ressuscitar" contratos cancelados ou corrigir status incorretos.
    """
    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        if not arquivo:
            flash("Selecione um arquivo Excel para processar.", "error")
            return redirect(url_for("sobrepor_status_view"))

        if not arquivo.filename.lower().endswith((".xls", ".xlsx")):
            flash("Envie um arquivo Excel (.xls ou .xlsx).", "error")
            return redirect(url_for("sobrepor_status_view"))

        try:
            resumo, mensagens = sobrepor_status_de_planilha(arquivo)
            flash(resumo, "success")
            for msg in mensagens[:50]:  # Limita a 50 mensagens
                flash(msg, "info")
            if len(mensagens) > 50:
                flash(f"... e mais {len(mensagens) - 50} registros.", "info")
        except Exception as e:
            logging.error(f"Erro ao sobrepor status via upload: {e}", exc_info=True)
            flash(f"Erro ao sobrepor status: {e}", "error")

        return redirect(url_for("sobrepor_status_view"))

    return render_template("sobrepor_status.html")


if __name__ == '__main__':
    import sys
    # Detecta se está rodando como executável (PyInstaller)
    if getattr(sys, 'frozen', False):
        # Rodando como executável - usar Waitress para produção
        print("=" * 60)
        print("iniciando sistema Bree com Waitress...")
        print("=" * 60)
        try:
            from waitress import serve
            print("INFO:waitress: Serving on http://0.0.0.0:5000")
            print("=" * 60)
            print("Sistema acessível em:")
            print("  - Local: http://localhost:5000")
            print("  - Rede: http://SEU_IP_LOCAL:5000")
            print("=" * 60)
            print("Para descobrir seu IP: execute 'ipconfig' no CMD")
            print("Pressione Ctrl+C para parar o servidor")
            print("=" * 60)
            serve(app, host='0.0.0.0', port=5000)
        except ImportError:
            print("ERRO: waitress não instalado!")
            print("Execute: pip install waitress")
            input("Pressione Enter para sair...")
            sys.exit(1)
        except Exception as e:
            print(f"ERRO ao iniciar servidor: {e}")
            input("Pressione Enter para sair...")
            sys.exit(1)
    else:
        # Rodando como script Python - modo desenvolvimento
        app.run(debug=True, host='0.0.0.0', port=5000)
